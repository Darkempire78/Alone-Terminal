import pandas as pd
import ast
import json
import codecs
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

def isNaN(num):
    return num != num

def excelExtractor(mobileDevice=False):
    # Read excel file into pandas DataFrames
    event_df = pd.read_excel('event_table.xlsx', sheet_name=0)
    add_event_df = pd.read_excel('event_table.xlsx', sheet_name=1)
    loot_df = pd.read_excel('event_table.xlsx', sheet_name=2)

    # For images
    event_df_images = load_workbook('event_table.xlsx').worksheets[0]
    event_df_image_loader = SheetImageLoader(event_df_images)

    event = {}
    
    # Iterate over rows in the event sheet
    index = 0
    for _, row in event_df.iterrows():
        index += 1
        # Ignore first line
        if row[0] == "zone":
            continue
        zone = row[0]
        type_ = row[1]

        if zone not in event:
            event[zone] = {"script": {}, "random": {}}

        name = row[2]
        image = row[3]

        cell = event_df_images.cell(row=index, column=4).coordinate
        if event_df_image_loader.image_in(cell):
            image = event_df_image_loader.get(cell)
            print(image)
            image.show()


        text = row[4]
        
        event[zone][type_][name] = {"text": text, "actions": {}}
        
        # Action 1
        event[zone][type_][name]["actions"]["action1"] = {
            "description": row[5] if isNaN(row[5]) == False else "<empty>",
            "next_message": row[6] if isNaN(row[6]) == False else "<empty>",
            "health": int(row[7]),
            "oxygen": int(row[8]),
            "loot": row[9].split(", ") if pd.notna(row[9]) else [],
            "add": ast.literal_eval(row[10]) if pd.notna(row[10]) else [],
            "remove": ast.literal_eval(row[11]) if pd.notna(row[11]) else [],
            "script": row[12].split(", ") if pd.notna(row[12]) else [],
            "zone": row[13] if isNaN(row[13]) == False else None,
            "require": row[14].split(", ") if pd.notna(row[14]) else []
        }

        # Action 2 if it exists
        if len(row) > 15 and pd.notna(row[14]):
            event[zone][type_][name]["actions"]["action2"] = {
                "description": row[15] if isNaN(row[15]) == False else "<empty>",
                "next_message": row[16] if isNaN(row[16]) == False else "<empty>",
                "health": int(row[17]),
                "oxygen": int(row[18]),
                "loot": row[19].split(", ") if pd.notna(row[19]) else [],
                "add": ast.literal_eval(row[20]) if pd.notna(row[20]) else [],
                "remove": ast.literal_eval(row[21]) if pd.notna(row[21]) else [],
                "script": row[22].split(", ") if pd.notna(row[22]) else [],
                "zone": row[23] if isNaN(row[23]) == False else None,
                "require": row[24].split(", ") if pd.notna(row[24]) else []
            }

    if mobileDevice:
        # Write the addEvent dictionary to a JSON file
        with codecs.open("data_for_mobile/Event.json", 'w', "utf-8") as file:
            file.write(json.dumps(event, indent=4, ensure_ascii=False))
    else:
        # Write the event dictionary to a python file
        with codecs.open("data/Event.py", 'w', "utf-8") as file:
            data = json.dumps(event, indent=4, ensure_ascii=False)
            data = data.replace("null", "None")
            file.write("event = " + data)

    # Process Loot sheet
    loot = {}
    
    for _, row in loot_df.iterrows():
        code_name = row[0]
        loot[code_name] = {
            "name": row[1],
            "health": int(row[2]),
            "oxygen": int(row[3]),
            "maxHealth": int(row[4]),
            "maxOxygen": int(row[5])
        }

    if mobileDevice:
        # Write the addEvent dictionary to a JSON file
        with codecs.open("data_for_mobile/Loot.json", 'w', "utf-8") as file:
            file.write(json.dumps(loot, indent=4, ensure_ascii=False))
    else:
        # Write the loot dictionary to a python file
        with codecs.open("data/Loot.py", 'w', "utf-8") as file:
            data = json.dumps(loot, indent=4, ensure_ascii=False)
            data = data.replace("null", "None")
            file.write("loot = " + data)

    # Process Add Event sheet
    addEvent = {}

    for _, row in add_event_df.iterrows():
        # Ignore first line
        if row[0] == "zone":
            continue

        name = row[2]
        addEvent[name] = {"text": row[3], "actions": {}}
        
        # Action 1
        addEvent[name]["actions"]["action1"] = {
            "description": row[4] if isNaN(row[4]) == False else "<empty>",
            "next_message": row[5] if isNaN(row[5]) == False else "<empty>",
            "health": int(row[6]),
            "oxygen": int(row[7]),
            "loot": row[8].split(", ") if pd.notna(row[8]) else [],
            "add": ast.literal_eval(row[9]) if pd.notna(row[9]) else [],
            "remove": ast.literal_eval(row[10]) if pd.notna(row[10]) else [],
            "script": row[11].split(", ") if pd.notna(row[11]) else [],
            "zone": row[12] if isNaN(row[12]) == False else None,
            "require": row[13].split(", ") if pd.notna(row[13]) else []
        }

        # Action 2 if it exists
        if len(row) > 14 and pd.notna(row[14]):
            addEvent[name]["actions"]["action2"] = {
                "description": row[14] if isNaN(row[14]) == False else "<empty>",
                "next_message": row[15] if isNaN(row[15]) == False else "<empty>",
                "health": int(row[16]),
                "oxygen": int(row[17]),
                "loot": row[18].split(", ") if pd.notna(row[18]) else [],
                "add": ast.literal_eval(row[19]) if pd.notna(row[19]) else [],
                "remove": ast.literal_eval(row[20]) if pd.notna(row[20]) else [],
                "script": row[21].split(", ") if pd.notna(row[21]) else [],
                "zone": row[22] if isNaN(row[22]) == False else None,
                "require": row[23].split(", ") if pd.notna(row[23]) else []
            }

    if mobileDevice:
        # Write the addEvent dictionary to a JSON file
        with codecs.open("data_for_mobile/AddEvent.json", 'w', "utf-8") as file:
            file.write(json.dumps(addEvent, indent=4, ensure_ascii=False))
    else:
        # Write the addEvent dictionary to a PYTHON file
        with codecs.open("data/AddEvent.py", 'w', "utf-8") as file:
            data = json.dumps(addEvent, indent=4, ensure_ascii=False)
            data = data.replace("null", "None")
            file.write("addEvent = " + data)
# Run the function
excelExtractor()
